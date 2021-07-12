# -*- coding: utf-8 -*-
"""
This is code for plotting results of GPC
The data is copied from corresonding excel file and is used to plot molar mass vs
cumullative distribution
"""
import openpyxl as op
import numpy as np
import matplotlib.pyplot as plt
import os


i=0
sample_dict = {1:'B1a', 2:'B1b', 3:'B2a', 4:'B2b', 5:'B3a', 6:'B3b', 7:'B4a',
                8:'B5a', 9:'Na1', 10:'Na1F', 11:'Na2', 12:'Na2F', 13:'W2', 14:'W2F',
                15:'W1F', 16:'W3F'} 
wb_1611 = op.load_workbook("../GPC/20201116_GPC/20201116_GPC.xlsx")
wb_0610 = op.load_workbook("../GPC/20201006_GPC/20201006_GPC_Results.xlsx")
wb_0812 = op.load_workbook("../GPC/20201208_GPC/20201208_GPC.xlsx")


# These plot points for 16/11 file
all_plot_points_1 = {}  
for active_sheet in wb_1611.worksheets:
    x_coordinate = []
    y_coordinate = []
    inner_dict = {}
    i += 1
    sample = sample_dict[i]    
# x and y coordinates are initilised here
# If the x and y cordinates needs to be changed just adjust the column in 
# append code below. For the row for loop the first number corresonds to starting 
# row in excel doc
# Column B= Retention time; column C= Molar Mass; column D=log molar mass;
# Column G= cumm dist rid; column H= vwd ELU ; column J= cumm dist vwd
    
    for row in range(49, active_sheet.max_row + 1):
        x_coordinate.append(active_sheet['C' + str(row)].value)
        y_coordinate.append(active_sheet['H' + str(row)].value)
        
        
        # code below print out all the graphs individually
    # figure_title = 'Sample {}'.format(sample) 
    # figure_name = 'Sample_{}'.format(sample) + '.png'
    # plt.figure(i, dpi=600), plt.clf()
    # plt.plot(x_coordinate, y_coordinate)
    # plt.title(figure_title)
    # plt.xlabel('molar mass')
    # plt.ylabel('VWD/MMD')
    # plt.grid()
    # plt.savefig(figure_name, bbox_inches='tight', dpi=600)
    
    inner_dict['x'] = x_coordinate 
    inner_dict['y'] = y_coordinate
    all_plot_points_1[i] = inner_dict 



# These plot points for 06/10 file
i=0
all_plot_points_2 = {}  
for active_sheet in wb_0610.worksheets:
    x_coordinate2 = []
    y_coordinate2 = []
    inner_dict2 = {}
    i += 1
    # sample = sample_dict[i]    
# x and y coordinates are initilised here
# If the x and y cordinates needs to be changed just adjust the column in 
# append code below. For the row for loop the first number corresonds to starting 
# row in excel doc
# Column B= Retention time; column C= Molar Mass; column D=log molar mass;
# Column G= cumm dist rid; column J= cumm dist vwd
    
    for row in range(49, active_sheet.max_row + 1):
        x_coordinate2.append(active_sheet['C' + str(row)].value)
        y_coordinate2.append(active_sheet['H' + str(row)].value)
        
    
    inner_dict2['x'] = x_coordinate2
    inner_dict2['y'] = y_coordinate2
    all_plot_points_2[i] = inner_dict2



# These points are for 08/12 file
i=0
all_plot_points_3 = {}  
for active_sheet in wb_0610.worksheets:
    x_coordinate3 = []
    y_coordinate3 = []
    inner_dict3 = {}
    i += 1
    # sample = sample_dict[i]    
# x and y coordinates are initilised here
# If the x and y cordinates needs to be changed just adjust the column in 
# append code below. For the row for loop the first number corresonds to starting 
# row in excel doc
# Column B= Retention time; column C= Molar Mass; column D=log molar mass;
# Column G= cumm dist rid; column J= cumm dist vwd
    
    for row in range(49, active_sheet.max_row + 1):
        x_coordinate3.append(active_sheet['C' + str(row)].value)
        y_coordinate3.append(active_sheet['H' + str(row)].value)
        
    
    inner_dict3['x'] = x_coordinate3
    inner_dict3['y'] = y_coordinate3
    all_plot_points_3[i] = inner_dict3



    
# Lignin distribution coeff plotting
# print(all_plot_points[1]['y'][1])
# stage1 = np.divide(all_plot_points_2[9]['y'][318:1852], all_plot_points_2[1]['y'][318:1852] )*38.36
stage1 = np.divide(all_plot_points_2[5]['y'][224:1852], all_plot_points_2[4]['y'][224:1852] )
batch = np.divide(all_plot_points_2[9]['y'][224:1852], all_plot_points_2[8]['y'][224:1852] )
# batch = (np.divide(all_plot_points_2[7]['y'][200:1852], all_plot_points_2[1]['y'][200:1852]) + 
          # np.divide(all_plot_points_2[9]['y'][200:1852], all_plot_points_2[1]['y'][200:1852]))*38.36




# plt.figure('Distribution Coeff', dpi=600), plt.clf()
# plt.ylim(0,3)
# plt.xlim(1000,10000)
# # plt.plot(all_plot_points_2[1]['x'][318:1852], stage1,'k-')
# # plt.plot(all_plot_points_2[1]['x'][318:1852], stage2,'-')
# plt.plot(all_plot_points_2[1]['x'][224:1852], batch, 'r:')
# # plt.plot(all_plot_points_2[1]['x'][224:1852], stage1, 'b--')
# plt.grid()
# # plt.title('Lignin Recovery')
# plt.xlabel('Molar Weight (g/mol)')
# plt.ylabel('Distribution Coefficient (-)')
# # plt.legend([' Batch LLX', 'Feed 1: CCS LLX'])




#  Comparison plot below
plt.figure('Compare', dpi=600), plt.clf()
plt.xlim(900,10000)
plt.ylim(0,10)
plt.plot(all_plot_points_1[4]['x'], all_plot_points_1[4]['y'], '--r' )
plt.plot(all_plot_points_2[7]['x'], all_plot_points_2[7]['y'], '-g' )
# plt.plot(all_plot_points_2[8]['x'], all_plot_points_2[8]['y'], 'g--' )
# plt.plot(all_plot_points_1[13]['x'], all_plot_points_1[13]['y'], 'r--' )
# plt.plot(all_plot_points_1[10]['x'], all_plot_points_1[10]['y'], 'k--' )
plt.grid()
# plt.title('Lignin Distribution')
plt.xlabel('Molar mass (gm/mol)')
plt.ylabel('VWD Singal(-)')
plt.legend([ 'Dirty Water', 'Extract (2-MTHF stream)'])
    
# Plotting for files from 08/12

# plt.figure('Compare', dpi=600), plt.clf()
# plt.xlim(500, 15000)
# plt.ylim(0,10)    
# # plt.plot(all_plot_points_3[2]['x'], all_plot_points_3[2]['y'])
# plt.plot(all_plot_points_3[5]['x'], all_plot_points_3[5]['y'])
# plt.plot(all_plot_points_3[7]['x'], all_plot_points_3[7]['y'])
# plt.plot(all_plot_points_3[8]['x'], all_plot_points_3[8]['y'])
# plt.plot(all_plot_points_3[9]['x'], all_plot_points_3[9]['y'])
# plt.grid()
# plt.title('Lignin Distribution')
# plt.xlabel('Molar mass (g/mol)')
# plt.ylabel('VWD Singal')
# plt.legend(['B1E', 'c1a','c2a','c3a'])



